from openpyxl import load_workbook
import sys
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import *
from PySide6.QtCore import QFile, QIODevice
from PySide6.QtGui import QBrush, QColor

TabelExcel = None
window = None
linii = 0
coloane = 0

def iesire():
    print("Am iesit")
    sys.exit(0)

def DespreNebun():
    print("Afisam")
    window.findChild(QPushButton, "pushButton_Despre").setText("Nu stiu sa fac popup sry")

def filtruCrazy():
    print("Filtram o nebunie mare")
    global window, TabelExcel
    Tabelul = window.findChild(QTableView, "tableWidget")
    Tabelul.setRowCount(0)
    #Tabelul.setColumnCount(0)
    #filtrul pentru anotimp
    copieTabel = []
    copieTabelLinii = 0
    elementeTabel = 0
    elementCautat = window.findChild(QComboBox, "comboBoxSezon").currentText()
    for i in range(1, linii-1):
        if elementCautat in TabelExcel[i-1][1] or elementCautat == "Any":
            Tabelul.insertRow(elementeTabel)
            elementeTabel += 1
            listaTemporara = []
            copieTabelLinii += 1
            for j in range(1, coloane):
                #Tabelul.setItem(elementeTabel-1, j-1, QTableWidgetItem(TabelExcel[i-1][j-1]))
                listaTemporara.append(str(TabelExcel[i-1][j-1]))
            copieTabel.append(listaTemporara)
    
    #filtru pentru vreme
    Tabelul.setRowCount(0)
    elementeTabel = 0
    elementCautat = window.findChild(QComboBox, "comboBoxVreme").currentText()
    for i in range(0, copieTabelLinii):
        if elementCautat in copieTabel[i][2] or elementCautat == "Any" or copieTabel[i][2] == "Any":
            Tabelul.insertRow(elementeTabel)
            elementeTabel += 1
            for j in range(1, coloane):
                Tabelul.setItem(elementeTabel-1, j-1, QTableWidgetItem(copieTabel[i][j-1]))
                if j-1 == 7:
                    if copieTabel[i][j-1] == "Yes":
                        Tabelul.item(elementeTabel-1, j-1).setBackground(QBrush(QColor(0,127,0,255)))
                    else:
                        Tabelul.item(elementeTabel-1, j-1).setBackground(QBrush(QColor(127,0,0,255)))  

def apasareCelula(row, column):
    global window
    Tabelu = window.findChild(QTableWidget, "tableWidget")
    print("Mi-ai apasat celula "+str(row)+" "+str(column))
    #if column == 7 and Tabelu.item(row, column).currentText() == "Yes":
    #    Tabelu.item(row, column).setBackground(QBrush(QColor(0,127,0,255)))


def IncarcaTabel():
    print("Incarc tabelul")
    global TabelExcel
    global window
    global coloane, linii
    FisieruExcel = load_workbook(filename = 'Fish_table.xlsx')
    WorksheetExcel = FisieruExcel.active
    print( WorksheetExcel.title)

    coloane = 1
    linii = 1
    while WorksheetExcel.cell(1,coloane).value:
         print("valuare: "+str(WorksheetExcel.cell(1,coloane).value)+"'")
         coloane += 1
    while WorksheetExcel.cell(linii,1).value:
         linii += 1
    #TabelExcel = [[""]*(coloane-1)]*(linii-1)
    TabelExcel = []

    Tabelul = window.findChild(QTableView, "tableWidget")
    Tabelul.setRowCount(linii-2)
    Tabelul.setColumnCount(coloane-1)

    for i in range(1, linii):
        listaRand = []
        for j in range(1, coloane):
            if i-1 > 0:
                celula =WorksheetExcel.cell(i, j).value
                #TabelExcel[i-2][j-1]
                listaRand.append(str(celula))
                Tabelul.setItem(i-2, j-1, QTableWidgetItem(str(celula)))
                if j-1 == 7:
                    if str(celula) == "Yes":
                        Tabelul.item(i-2, j-1).setBackground(QBrush(QColor(0,127,0,255)))
                    else:
                        Tabelul.item(i-2, j-1).setBackground(QBrush(QColor(127,0,0,255)))
            else:
                Tabelul.setHorizontalHeaderItem(j-1, QTableWidgetItem(WorksheetExcel.cell(i, j).value))
        if listaRand:
            TabelExcel.append(listaRand)
    Tabelul.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    print("Am incarcat")

if __name__ == "__main__":
    app = QApplication(sys.argv)

    ui_file_name = "Interfata.ui"
    ui_file = QFile(ui_file_name)
    if not ui_file.open(QIODevice.ReadOnly):
        print(f"Cannot open {ui_file_name}: {ui_file.errorString()}")
        sys.exit(-1)
    loader = QUiLoader()
    window = loader.load(ui_file)
    ui_file.close()

    ButonIesire = window.findChild(QPushButton, "pushButton_Iesire")
    ButonIesire.clicked.connect(iesire)
    ButonIncarca = window.findChild(QPushButton, "pushButton_Incarca")
    ButonIncarca.clicked.connect(IncarcaTabel)
    ButonDespre = window.findChild(QPushButton, "pushButton_Despre")
    ButonDespre.clicked.connect(DespreNebun)
    FiltruCombo = window.findChild(QComboBox, "comboBoxSezon")
    FiltruCombo.currentIndexChanged.connect(filtruCrazy)
    FiltruCombo2 = window.findChild(QComboBox, "comboBoxVreme")
    FiltruCombo2.currentIndexChanged.connect(filtruCrazy)
    TabelMare = window.findChild(QTableWidget, "tableWidget").cellClicked.connect(apasareCelula)
    if not window:
        print(loader.errorString())
        sys.exit(-1)
    window.show()

    sys.exit(app.exec())